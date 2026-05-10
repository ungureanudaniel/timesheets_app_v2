from django.views.generic import FormView, TemplateView
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from django.urls import reverse_lazy
from timesheet.models import Timesheet
from .forms import ReportPeriodForm
from datetime import timedelta, datetime
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
from users.models import CustomUser
import openpyxl
from openpyxl.styles import Font, Alignment
import calendar
from django.utils.translation import gettext as _

User = CustomUser

class ReportGeneratorView(LoginRequiredMixin, FormView):
    template_name = 'reports/generate_report.html'
    form_class = ReportPeriodForm
    success_url = reverse_lazy('report_results')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # This is the "user" that __init__ pops
        return kwargs

    def form_valid(self, form):
        period = form.cleaned_data['period']
        custom_start = form.cleaned_data.get('custom_start_date')
        custom_end = form.cleaned_data.get('custom_end_date')
        selected_user = form.cleaned_data.get('user')

        if selected_user:
            target_user_id = selected_user.id
        else:
            target_user_id = self.request.user.id

        # Calculate date range
        today = timezone.now().date()
        start_date = today
        end_date = today

        if period == 'current_week':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == 'last_week':
            start_date = today - timedelta(days=today.weekday() + 7)
            end_date = start_date + timedelta(days=6)
        elif period == 'current_month':
            start_date = today.replace(day=1)
            end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        elif period == 'last_month':
            start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
            end_date = start_date + timedelta(days=calendar.monthrange(start_date.year, start_date.month)[1] - 1)
        elif period == 'last_year':
            start_date = today.replace(year=today.year - 1, month=1, day=1)
            end_date = today.replace(year=today.year - 1, month=12, day=31)
        elif period == 'current_year':
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
        elif period == 'custom' and (not custom_start or not custom_end):
            form.add_error(None, _("Please provide both start and end dates for the custom range."))
            return self.form_invalid(form)
        elif custom_start and custom_end:
            start_date = custom_start
            end_date = custom_end
        else:
            form.add_error('period', _("Invalid period selection."))
            return self.form_invalid(form)
        
        # Store in session for results view
        self.request.session['report_data'] = {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'period_label': str(dict(form.fields['period'].choices).get(period, period)),
            'user_ids': target_user_id if target_user_id else []
        }
        return super().form_valid(form)


class ReportResultsView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/report_results.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report_data = self.request.session.get('report_data', {})
        
        if not report_data:
            return context
        
        start_date = datetime.strptime(report_data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(report_data.get('end_date'), '%Y-%m-%d').date()

        session_user_id = report_data.get('user_id')
    
        if session_user_id:
            user_id = session_user_id
        else:
            user_id = self.request.user.pk

        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            report_user = User.objects.get(id=user_id)
            context['report_user'] = report_user
        except User.DoesNotExist:
            context['report_user'] = self.request.user
            user_id = self.request.user.pk

        timesheets = Timesheet.objects.filter(
            user_id=user_id,
            date__range=[start_date, end_date]
        ).select_related('activity', 'fundssource').annotate(
            img_nr=Count('timesheet_images')
        ).order_by('date', 'start_time')

        detailed_data = []
        activity_totals = {}
        grand_total_decimal = 0
        
        for ts in timesheets:
            hours_dec, hours_str = self._get_hour_data(ts)
            grand_total_decimal += hours_dec
            
            detailed_data.append({
                'id': ts.pk,
                'date': ts.date,
                'start_time': ts.start_time,
                'end_time': ts.end_time,
                'activity_code': ts.activity.code,
                'activity_name': ts.activity.name,
                'description': ts.description,
                'hours': hours_str,
                'image_count': getattr(ts, 'img_nr', 0),
            })

            code = ts.activity.code
            activity_totals[code] = activity_totals.get(code, 0) + hours_dec
        
        total_h = int(grand_total_decimal)
        total_m = int(round((grand_total_decimal - total_h) * 60))
        
        context.update({
            'report_title': _("Detailed Activity Report"),
            'period_detail': f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}",
            'report_data': detailed_data,
            'total_hours': f"{total_h}h {total_m:02d}m",
            'total_entries': timesheets.count(),
            'chart_labels': list(activity_totals.keys()),
            'chart_values': [float(v) for v in activity_totals.values()],
        })
        return context

    def _get_hour_data(self, timesheet):
        """
        Calculates duration and returns a tuple: (decimal_hours, display_string)
        Example: (8.5, "8h 30m")
        """
        if timesheet.start_time and timesheet.end_time:
            start = datetime.combine(datetime.today(), timesheet.start_time)
            end = datetime.combine(datetime.today(), timesheet.end_time)
            
            if end < start:
                end += timedelta(days=1)
            
            duration_seconds = (end - start).total_seconds()
            decimal_hours = duration_seconds / 3600
            day_of_week = timesheet.date.weekday() # Monday=0, Sunday=6

            if day_of_week <= 3:  # Monday-Thursday
                capped_decimal = min(decimal_hours, 8.5)
            elif day_of_week == 4:  # Friday
                friday_deadline = datetime.combine(timesheet.date, datetime.strptime("14:00", "%H:%M").time())     
                   
            h = int(decimal_hours)
            m = int(round((decimal_hours - h) * 60))
            display_string = f"{h}h {m:02d}m"
            
            return decimal_hours, display_string
        
        return 0, "0h 00m"
    
    def _generate_detailed_report(self, timesheets):
        detailed_data = []
        # Note the prefetch name here must match your model's related_name
        timesheets_with_imgs = timesheets.prefetch_related('timesheet_images').order_by('date', 'start_time')
        
        for ts in timesheets_with_imgs:
            hours, duration_display = self._get_hour_data(ts)
            detailed_data.append({
                'date': ts.date,
                'hours': hours,
                'duration_display': duration_display,
                'fundssource_name': ts.fundssource.name if ts.fundssource else '',
                'activity_code': ts.activity.code,
                'description': ts.description or '',
                'start_time': ts.start_time,
                'end_time': ts.end_time,
                # Use 'timesheet_images' if that is your related_name
                'images': ts.timesheet_images.all() 
            })
        return detailed_data

from PIL import Image
from io import BytesIO

def process_image_for_pdf(image_field, max_size=(800, 800)):
    """
    Takes an ImageField, resizes it, compresses it, 
    and returns a BytesIO object for the PDF.
    """
    # Open the image using Pillow
    img = Image.open(image_field)
    
    # Convert to RGB if it's RGBA (prevents errors with JPEGs)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    
    # 1. Resize: Maintain aspect ratio
    img.thumbnail(max_size, Image.Resampling.LANCZOS)
    
    # 2. Compress: Save to a buffer with reduced quality
    output_buffer = BytesIO()
    img.save(output_buffer, format="JPEG", quality=70, optimize=True)
    output_buffer.seek(0)
    
    return output_buffer

import os
from django.conf import settings
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from io import BytesIO
from reportlab.platypus import Image, Spacer
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
FONT_DIR = os.path.join(settings.BASE_DIR, 'static', 'fonts/dejavu')

try:
    pdfmetrics.registerFont(TTFont('DejaVu', os.path.join(FONT_DIR, 'dejavu-sans-webfont.ttf')))
    pdfmetrics.registerFont(TTFont('DejaVu-Bold', os.path.join(FONT_DIR, 'dejavu-sans-bold.ttf')))
    FONT_NAME = 'DejaVu'
    BOLD_FONT = 'DejaVu-Bold'
except Exception as e:
    # Fallback if you forgot to download the files
    print(f"!!! FONT ERROR: {e}")
    FONT_NAME = 'Helvetica'
    BOLD_FONT = 'Helvetica-Bold'

class ExportPDFView(LoginRequiredMixin, TemplateView):
    def _format_hours_to_hm(self, decimal_hours):
        """Converts 8.5 to '8h 30m'"""
        hours = int(decimal_hours)
        minutes = int(round((decimal_hours - hours) * 60))
        return f"{hours}h {minutes:02d}m"

    def get(self, request, *args, **kwargs):
        report_data = request.session.get('report_data', {})
        if not report_data:
            return HttpResponse("No report data found")

        buffer = BytesIO()
        # Adjusted margins to fit content better
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # --- 1. DATA FETCHING ---
        user_id = report_data.get('user_id')
        start_date = datetime.strptime(report_data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(report_data.get('end_date'), '%Y-%m-%d').date()

        # Update styles for Unicode
        styles['Title'].fontName = BOLD_FONT
        styles['Normal'].fontName = FONT_NAME
        styles['Heading2'].fontName = BOLD_FONT
        styles['Heading3'].fontName = BOLD_FONT

        # --- 2. PRE-CALCULATE TOTALS (One Loop Only) ---
        timesheets = Timesheet.objects.filter(
            user_id=user_id,
            date__range=[start_date, end_date]
        ).select_related('activity', 'fundssource').annotate(
            img_nr=Count('timesheet_images')
        ).order_by('date', 'start_time')

        total_hours_decimal = 0
        activity_totals = {}
        for ts in timesheets:
            h = self._calculate_hours(ts)
            total_hours_decimal += h
            code = ts.activity.code
            activity_totals[code] = activity_totals.get(code, 0) + h

        # --- 3. HEADER & SUMMARY TABLE ---
        elements.append(Paragraph(f"Raport de activitate", styles['Title']))
        elements.append(Paragraph(f"Perioadă: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        standard_day = 8.5
        total_days = total_hours_decimal / standard_day
        
        summary_table_data = [
            [_('Total Time'), _('Work Days (8h 30m)'), _('Entries')],
            [
                self._format_hours_to_hm(total_hours_decimal), 
                f"{total_days:.2f}", 
                str(timesheets.count())
            ]
        ]
        
        st = Table(summary_table_data, colWidths=[2.5 * inch] * 3)
        st.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(st)
        elements.append(Spacer(1, 0.3 * inch))

        # --- 4. BAR CHART ---
        if activity_totals:
            chart_buffer = self._generate_bar_chart(activity_totals)
            chart_img = Image(chart_buffer)
            chart_img.drawHeight = 3.0 * inch
            chart_img.drawWidth = 3.0 * inch 
            elements.append(chart_img)
            elements.append(Spacer(1, 0.3 * inch))

        # --- 5. DETAILED LOG LOOP ---
        elements.append(Paragraph(_("Detailed Log"), styles['Heading2']))
        current_date = None
        
        for ts in timesheets:
            if ts.date != current_date:
                current_date = ts.date
                elements.append(Spacer(1, 0.1 * inch))
                date_style = styles['Heading3']
                date_style.backColor = colors.whitesmoke
                elements.append(Paragraph(f"Data: {current_date.strftime('%d.%m.%Y')}", date_style))
                elements.append(Spacer(1, 0.05 * inch))

            h = self._calculate_hours(ts)
            entry_data = [
                [_('Time'), _('Activity'), _('Source'), _('Hrs'), _('Description')],
                [
                    f"{ts.start_time.strftime('%H:%M')}-{ts.end_time.strftime('%H:%M')}",
                    Paragraph(f"<b>{ts.activity.code}</b><br/>{ts.activity.name}", styles['Normal']),
                    ts.fundssource.name if ts.fundssource else '-',
                    self._format_hours_to_hm(h), 
                    Paragraph(ts.description or '', styles['Normal'])
                ]
            ]
            
            t = Table(entry_data, colWidths=[0.8*inch, 1.8*inch, 1.0*inch, 0.7*inch, 3.0*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t)

            # Images
            ts_images = ts.timesheet_images.all()
            if ts_images:
                img_list = []
                for img_obj in ts_images:
                    try:
                        img = Image(img_obj.image.path)
                        aspect = img.imageHeight / float(img.imageWidth)
                        img.drawHeight = 1.2 * inch
                        img.drawWidth = (1.2 * inch) / aspect
                        img_list.append(img)
                    except: continue
                
                if img_list:
                    rows = [img_list[i:i + 3] for i in range(0, len(img_list), 3)]
                    img_table = Table(rows, hAlign='CENTER')
                    wrapper = Table([[img_table]], colWidths=[7.3 * inch])
                    wrapper.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('TOPPADDING', (0, 0), (-1, -1), 5),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
                    ]))
                    elements.append(wrapper)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="activity_report.pdf"'
        return response

    def _generate_pie_chart(self, data_dict):
        plt.figure(figsize=(6, 6)) # Keep it square to prevent vertical distortion
        
        # Check for diacritics in labels
        labels = list(data_dict.keys())
        values = list(data_dict.values())
        
        # Try to set a font that exists on Windows or Linux
        # Windows: Arial, Linux: DejaVu Sans
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'sans-serif']
        plt.rcParams['font.family'] = 'sans-serif'

        plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=140)
        plt.axis('equal') 
        
        chart_buffer = BytesIO()
        plt.savefig(chart_buffer, format='png', bbox_inches='tight')
        plt.close()
        chart_buffer.seek(0)
        return chart_buffer

    def _generate_bar_chart(self, data_dict):
        # Filter data to remove zeros (makes the chart cleaner)
        clean_data = {k: v for k, v in data_dict.items() if v > 0}
        
        if not clean_data:
            # Fallback if no data exists to prevent Matplotlib crash
            clean_data = {"No Data": 0}

        labels = list(clean_data.keys())
        values = list(clean_data.values())

        # Adjust figure height based on the number of items
        fig_height = max(4, len(labels) * 0.5)
        plt.figure(figsize=(8, fig_height))
        
        # Font setup for diacritics
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'sans-serif']
        plt.rcParams['font.family'] = 'sans-serif'

        # Create horizontal bars
        bars = plt.barh(labels, values, color='#007bff')

        # Add labels to the end of each bar for clarity
        for bar in bars:
            width = bar.get_width()
            plt.text(width + 0.1, bar.get_y() + bar.get_height()/2, 
                    f'{width}h', va='center', fontsize=10)

        plt.xlabel('Hours Worked')
        plt.title('Work Distribution')
        
        # Improve layout so labels aren't cut off
        plt.tight_layout()

        # Save to memory buffer
        chart_buffer = BytesIO()
        plt.savefig(chart_buffer, format='png', bbox_inches='tight', dpi=300)
        plt.close()
        chart_buffer.seek(0)
        
        return chart_buffer

    def _calculate_hours(self, timesheet):
        if timesheet.start_time and timesheet.end_time:
            start = datetime.combine(datetime.today(), timesheet.start_time)
            end = datetime.combine(datetime.today(), timesheet.end_time)
            if end < start: end += timedelta(days=1)
            return (end - start).total_seconds() / 3600
        return 0

class ExportExcelView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        report_data = request.session.get('report_data', {})
        
        if not report_data:
            return HttpResponse(_("No report data found"))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="activity_report.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title=_("Activity Report"))
        else:
            ws.title = _("Activity Report")
        
        # Get dates from report data
        start_date = datetime.strptime(report_data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(report_data.get('end_date'), '%Y-%m-%d').date()
        
        # Get report data
        timesheets = Timesheet.objects.filter(
            user_id=report_data.get('user_id'),
            date__range=[start_date, end_date]
        ).select_related('activity', 'fundssource')
            
        headers = ['Date', 'Start Time', 'End Time', 'Activity Code', 'Activity Name', 'Funds Source', 'Hours Worked', 'Description']
        ws.append(headers)
        for timesheet in timesheets.order_by('date'):
            hours = self._calculate_hours(timesheet)
                
            ws.append([
                    timesheet.date,
                    timesheet.start_time.strftime('%H:%M') if timesheet.start_time else '',
                    timesheet.end_time.strftime('%H:%M') if timesheet.end_time else '',
                    timesheet.activity.code,
                    timesheet.activity.name,
                    timesheet.fundssource.name if timesheet.fundssource else '',
                    round(hours, 2),
                    timesheet.description or ''
                ])
        # Style the header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        wb.save(response)
        return response
    
    def _calculate_hours(self, timesheet):
        """Calculate hours from start_time and end_time (both are time objects)"""
        if timesheet.start_time and timesheet.end_time:
            # Convert time objects to datetime objects for the same day to calculate difference
            start_datetime = datetime.combine(datetime.today(), timesheet.start_time)
            end_datetime = datetime.combine(datetime.today(), timesheet.end_time)
            
            # Handle case where end_time is after midnight (next day)
            if end_datetime < start_datetime:
                end_datetime = end_datetime + timedelta(days=1)
            
            duration = end_datetime - start_datetime
            return duration.total_seconds() / 3600
        return 0

