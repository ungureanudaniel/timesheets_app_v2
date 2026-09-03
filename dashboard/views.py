import calendar

from django.conf.locale import da
from django.shortcuts import render, redirect
from django.http import JsonResponse, request
from django.core.paginator import Paginator
from django.urls import reverse_lazy
from django.utils import timezone

from timesheets_main import settings
from .forms import PALActivitiesUploadForm, FundsSourceForm, PALActivityForm
from django.db.models import Count, Prefetch, Sum, F, ExpressionWrapper, fields, FloatField, Q
from django.contrib.auth import get_user_model
from dashboard.forms import ActivityProgramForm
from dashboard.models import ActivityProgram
from .utils import format_minutes
from timesheet.models import Activity, FundsSource
from users.models import CustomUser
from timesheet.models import Timesheet
from io import BytesIO
from django.views import View
from django.views.generic import CreateView, ListView, TemplateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from natsort import natsorted
import openpyxl
from django.views.generic import TemplateView
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from datetime import date, datetime, timedelta
from django.core.mail import send_mail
from django.db.models import Count
from timesheet.models import Timesheet
import holidays
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

def automated_task_runner(request):
    # Security check: Only let the pinger in
    secret_key = settings.TASK_RUNNER_KEY
    if request.GET.get('key') != secret_key:
        return HttpResponseForbidden("Invalid Key")

    task = request.GET.get('task')
    today = timezone.now().date()

    if task == "monday_summary":
        # 1. Send Office Summary + 2. Weekly Reminder to Reporters
        send_office_weekly_summary(today)
        return HttpResponse("Monday tasks completed")

    elif task == "friday_reminder":
        # 3. Friday Afternoon Reminder
        send_reporter_reminders("Friday Reminder: Please finish your reports before the weekend!")
        return HttpResponse("Friday reminders sent")

    elif task == "monthly_report":
        # 4. 1st of the Month Reminder
        send_reporter_reminders("Monthly Reminder: It is the 1st of the month. Please finalize last month's report.")
        return HttpResponse("Monthly reminders sent")

    return HttpResponse("No task specified")

# Helper functions to keep it clean
def send_office_weekly_summary(today):
    last_week = today - timedelta(days=7)
    summary = Timesheet.objects.filter(date__range=[last_week, today - timedelta(days=1)])\
        .values('user__username').annotate(days=Count('date', distinct=True))
    
    body = "Weekly Audit:\n" + "\n".join([f"{s['user__username']}: {s['days']} days" for s in summary])
    send_mail("Weekly Summary", body, "system@company.com", ["office@company.com"])

def send_reporter_reminders(msg):
    from django.contrib.auth.models import User
    emails = User.objects.filter(is_staff=False).values_list('email', flat=True)
    send_mail("Report Reminder", msg, "system@company.com", list(emails))

def sanitize_romanian(text):
    if not text:
        return ""
    replacements = {
        'ă': 'a', 'Ă': 'A',
        'ș': 's', 'Ș': 'S',
        'ț': 't', 'Ț': 'T',
        'â': 'a', 'Â': 'A',
        'î': 'i', 'Î': 'I'
    }
    for char, rep in replacements.items():
        text = text.replace(char, rep)
    return text

# main admin dashboard view.
def dashboard(request):
    template = "dashboard/dashboard.html"

    context = {}
    return render(request, template, context)

# ==============Data analytics============
class AnalyticsView(ListView):
    template_name = "dashboard/analytics.html"

    queryset = CustomUser.objects.all()
    paginate_by = 20

    def get(self, request, **kwargs):
        # get each individual userprofile safely
        user_profile = getattr(self.request.user, 'customuser', None)
        # If no customuser attribute (AnonymousUser or profile not created), user_profile will be None
        print(user_profile)
        return super().get(request, **kwargs)
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

class PALActivitiesListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Activity
    template_name = 'dashboard/pal.html'
    context_object_name = 'activities'
    paginate_by = 10

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_queryset(self):
        qs = Activity.objects.all()
        return natsorted(qs, key=lambda x: x.code)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        queryset = self.get_queryset()
        
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['activities'] = page_obj
        context['object_list'] = page_obj
        context['page_obj'] = page_obj
        context['is_paginated'] = page_obj.has_other_pages()
        
        return context


User = get_user_model()

class HoursSummaryTableView(LoginRequiredMixin, TemplateView):
    template_name = "dashboard/hours_summary.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        
        # Parse selected period or default to current month
        period_query = request.GET.get('selected_period')
        if not period_query:
            current_date = datetime.now()
            period_query = current_date.strftime('%Y-%m')

        # Split into numerical items
        year, month = map(int, period_query.split('-'))

        # Instantiate Romanian Public Holidays rules engine for this target year
        ro_holidays = holidays.Romania(years=year)

        # Month names for display
        ro_months = {
            1: "Ianuarie", 2: "Februarie", 3: "Martie", 4: "Aprilie",
            5: "Mai", 6: "Iunie", 7: "Iulie", 8: "August",
            9: "Septembrie", 10: "Octombrie", 11: "Noiembrie", 12: "Decembrie"
        }
        ro_days_short = ["L", "M", "M", "J", "V", "S", "D"]

        # Compute structural day lists for selected calendar space
        num_days = calendar.monthrange(year, month)[1]
        month_days_list = []
        
        # Track valid baseline legal working days for the contract type norm math
        actual_working_days_count = 0
        
        for d in range(1, num_days + 1):
            current_date = date(year, month, d)
            weekday_index = calendar.weekday(year, month, d)
            
            is_weekend = weekday_index in [5, 6]
            is_holiday = current_date in ro_holidays
            holiday_name = ro_holidays.get(current_date, "") if is_holiday else ""

            if not is_weekend and not is_holiday:
                actual_working_days_count += 1

            month_days_list.append({
                'day_num': d,
                'day_letter': ro_days_short[weekday_index],
                'is_weekend': is_weekend,
                'is_holiday': is_holiday,
                'holiday_name': holiday_name
            })

        context['current_period'] = period_query
        context['current_month_year'] = f"{ro_months[month]} {year}"
        context['month_days'] = month_days_list

        if request.user.is_staff or request.user.groups.filter(name='Managers').exists():
            employees = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
        else:
            pass

        monthly_timesheets = Timesheet.objects.filter(
            date__year=year, 
            date__month=month
        ).select_related('activity')

        employees = employees.prefetch_related(
            Prefetch('timesheet_set', queryset=monthly_timesheets, to_attr='cached_month_timesheets')
        )

        employee_data = []

        for emp in employees:
            # Initialize every single day with a default structure
            days_matrix = {d: {'type': 'none', 'hours': ''} for d in range(1, num_days + 1)}
            
            total_hours_worked = 0.0
            total_minutes_worked = 0
            total_co_days = 0
            total_cm_days = 0
            worked_days_set = set() 
            co_days_set = set()
            cm_days_set = set()
            ef_days_set = set()
            # Loop through pre-fetched records smoothly
            cached_sheets = getattr(emp, 'cached_month_timesheets', [])
            for ts in cached_sheets:
                day_number = ts.date.day
                
                activity_code = ts.activity.code.upper() if (ts.activity and ts.activity.code) else ""
                activity_name = ts.activity.name.upper() if (ts.activity and ts.activity.name) else ""
                
                is_co = (
                    "CO" == activity_code or 
                    "ODIHNA" in activity_name or 
                    "ODIHNĂ" in activity_name or
                    "CONCEDIU DE ODIHNĂ" in activity_name or
                    "CONCEDIU ANUAL" in activity_name
                )
                
                is_cm = (
                    "CM" == activity_code or 
                    "MEDICAL" in activity_name or
                    "CONCEDIU MEDICAL" in activity_name or
                    "BOALA" in activity_name
                )
                
                is_ef = (
                    "EF" == activity_code or 
                    "EVENIMENT FAMILIAL" in activity_name or
                    "FAMILIAL" in activity_name
                )
                if is_co:
                    days_matrix[day_number] = {'type': 'CO', 'hours': 'CO'}
                    co_days_set.add(day_number)
                elif is_cm:
                    days_matrix[day_number] = {'type': 'CM', 'hours': 'CM'}
                    cm_days_set.add(day_number)
                elif is_ef:
                    days_matrix[day_number] = {'type': 'EF', 'hours': 'EF'}
                    ef_days_set.add(day_number)
                else:
                    # Calculate worked hours for this timesheet entry
                    if hasattr(ts, 'duration_decimal') and ts.duration_decimal is not None:
                        hours = float(ts.duration_decimal)
                    elif ts.start_time and ts.end_time:
                        today_dummy = datetime.today()
                        dt1 = datetime.combine(today_dummy, ts.start_time)
                        dt2 = datetime.combine(today_dummy, ts.end_time)
                        hours = max(0.0, (dt2 - dt1).total_seconds() / 3600.0)
                    else:
                        hours = 8.0
                    minutes = round(hours * 60)  
                    current_entry = days_matrix[day_number]
                    if current_entry['type'] == 'work':
                        existing_minutes = int(current_entry['hours']) if current_entry['hours'] else 0
                        new_total_minutes = existing_minutes + minutes
                        days_matrix[day_number]['hours'] = new_total_minutes  # Store as minutes
                    else:
                        days_matrix[day_number] = {'type': 'work', 'hours': minutes}

                    total_minutes_worked += minutes
                    total_hours_worked += hours
                    # Track this day as a worked day for meal ticket eligibility
                    worked_days_set.add(day_number)
            
            eligible_meal_ticket_days = worked_days_set - co_days_set - cm_days_set
            # Standard Romanian Norm setup subtracting statutory bank holidays 
            norma_hours = actual_working_days_count * 8
            norma_minutes = norma_hours * 60
            employee_data.append({
                'employee': emp,
                'norma_hours': norma_hours,
                'norma_minutes': norma_minutes,
                'days_matrix': days_matrix,  
                'total_hours_worked': round(total_hours_worked, 1),
                'total_minutes_worked': total_minutes_worked,
                'total_co_days': len(co_days_set),
                'total_cm_days': len(cm_days_set),
                'total_ef_days': len(ef_days_set),
                'meal_tickets_count': len(eligible_meal_ticket_days)
            })
        serializable_employee_data = []

        for emp_data in employee_data:
            emp = emp_data.get('employee')
            serializable_emp_data = {
                'employee_id': emp.id,
                'employee_name': f"{emp.first_name} {emp.last_name}".strip() if emp else "Unknown",
                'norma_hours': emp_data.get('norma_hours'),
                'norma_minutes': emp_data.get('norma_minutes'),
                'days_matrix': emp_data.get('days_matrix'),
                'total_hours_worked': emp_data.get('total_hours_worked'),
                'total_minutes_worked': emp_data.get('total_minutes_worked'),
                'total_co_days': emp_data.get('total_co_days'),
                'total_cm_days': emp_data.get('total_cm_days'),
                'total_ef_days': emp_data.get('total_ef_days'),
                'meal_tickets_count': emp_data.get('meal_tickets_count')
            }
            serializable_employee_data.append(serializable_emp_data)

        print(f"Storing {len(serializable_employee_data)} employees in session")
        if serializable_employee_data:
            print(f"First employee: {serializable_employee_data[0]}")


        request.session['pdf_employee_data'] = serializable_employee_data
        request.session['pdf_period'] = {
            'year': year,
            'month': month,
            'period_query': period_query,
        }
        request.session.modified = True
        print(f"Session keys after saving: {list(request.session.keys())}")
        print(f"Session data count: {len(request.session.get('pdf_employee_data', []))}")

        # context['employee_data'] = sorted(employee_data, key=lambda x: x['employee'].last_name)
        context['employee_data'] = employee_data
        context['current_period'] = period_query
        return context

class TimesheetPDFView(View):
    def get(self, request, *args, **kwargs):
        # 1. Parse target period (YYYY-MM)
        # Get data from session
        employee_data = request.session.get('pdf_employee_data', [])
        period_data = request.session.get('pdf_period', {})
        
        selected_period = request.GET.get('selected_period', datetime.now().strftime('%Y-%m'))
        if selected_period and not employee_data:
            try:
                year, month = map(int, selected_period.split('-'))
                return HttpResponse("Please load the summary page first.", status=400)
            except ValueError:
                pass
        if not employee_data:
            return HttpResponse("No data available. Please go back and load the summary first.", status=400)
        year = period_data.get('year', datetime.now().year)
        month = period_data.get('month', datetime.now().month)
        selected_period = period_data.get('period_query', f"{year}-{month:02d}")

        # Get total number of days in selected month as strict integers
        _, num_days = calendar.monthrange(year, month)
        month_days = list(range(1, num_days + 1))

        # 3. Setup PDF Document (A4 Landscape)
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=15,
            rightMargin=15,
            topMargin=15,
            bottomMargin=15
        )
        
        elements = []

        # 4. Setup Styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=12,
            leading=14,
            alignment=1,
            textColor=colors.HexColor('#1a252f')
        )
        
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=1,
            textColor=colors.HexColor('#555555')
        )

        header_cell_style = ParagraphStyle(
            'HeaderCell',
            parent=styles['Normal'],
            fontSize=6,
            leading=7,
            alignment=1,
            fontName='Helvetica-Bold'
        )

        name_cell_style = ParagraphStyle(
            'NameCell',
            parent=styles['Normal'],
            fontSize=6.5,
            leading=8,
            fontName='Helvetica-Bold'
        )

        body_cell_style = ParagraphStyle(
            'BodyCell',
            parent=styles['Normal'],
            fontSize=6,
            leading=7,
            alignment=1
        )

        sig_title_style = ParagraphStyle(
            'SigTitle',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            fontName='Helvetica-Bold',
            alignment=1
        )

        sig_name_style = ParagraphStyle(
            'SigName',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            alignment=1
        )

        # 5. Build Top Header Info
        month_name = calendar.month_name[month]
        elements.append(Paragraph(("FOAIE COLECTIVA DE PREZENTA"), title_style))
        elements.append(Paragraph(f"{('Pontaj lunar')} — {month_name} {year}", subtitle_style))
        elements.append(Spacer(1, 8))

        # 6. Construct Dynamic Table Headers (Using pure integer day numbers)
        row1 = [
            Paragraph("<b>Nr.<br/>crt.</b>", header_cell_style),
            Paragraph("<b>Nume Prenume</b>", header_cell_style),
            Paragraph("<b>Norma</b>", header_cell_style),
        ]
        
        for day_int in month_days:
            row1.append(Paragraph(f"<b>{day_int}</b>", header_cell_style))
            
        row1.extend([
            Paragraph("<b>Total<br/>ore</b>", header_cell_style),
            Paragraph("<b>Zile<br/>CO</b>", header_cell_style),
            Paragraph("<b>Zile<br/>CM</b>", header_cell_style),
            Paragraph("<b>Zile<br/>EF</b>", header_cell_style),
            Paragraph("<b>Tichete<br/>Masa</b>", header_cell_style),
        ])

        table_data = [row1]

        # 7. Populate Employee Rows (Strict Type Coercion to prevent 'Day' callables)
        for idx, row in enumerate(employee_data, start=1):
            # Safe name extraction
            if isinstance(row, dict):
                emp_name = row.get('employee_name', f"Angajat {idx}")
                norma = row.get('norma', row.get('norma_hours', 168))
                days_matrix = row.get('days_matrix', {})
                total_minutes = row.get('total_minutes_worked')
                total_formatted = format_minutes(total_minutes)
                co_days = row.get('co_days', row.get('total_co_days', 0))
                cm_days = row.get('cm_days', row.get('total_cm_days', 0))
                ef_days = row.get('ef_days', row.get('total_ef_days', 0))
                meal_tickets = row.get('meal_tickets', row.get('meal_tickets_count', 0))
            else:
                emp_name = getattr(row, 'employee_name')
                norma = getattr(row, 'norma', 168)
                days_matrix = getattr(row, 'days_matrix', {})
                total_minutes = getattr(row, 'total_minutes_worked', 0)
                total_formatted = format_minutes(total_minutes)
                co_days = getattr(row, 'co_days', 0)
                cm_days = getattr(row, 'cm_days', 0)
                ef_days = getattr(row, 'ef_days', 0)
                meal_tickets = getattr(row, 'meal_tickets', 0)

            # fetch and format employee name
            emp_name = f"{emp_name}".strip() if emp_name else f"Angajat {idx}"

            data_row = [
                Paragraph(str(idx), body_cell_style),
                Paragraph(emp_name, name_cell_style),
                Paragraph(str(norma), body_cell_style),
            ]
            if employee_data:
                        print(f"First employee days_matrix: {employee_data[0].get('days_matrix', {})}")
                        print(f"Printing day 11 data: {format_minutes(employee_data[0].get('days_matrix', {}).get('11', {}).get('hours', 'N/A'))}")
            # Populate matrix days
            for day_int in month_days:
                # Get the day's data
                day_data = days_matrix.get(str(day_int), {})
                
                # Debug: print what we got
                print(f"Row {idx}, Day {day_int}: day_data = {day_data.get('hours', 'N/A')}")
                
                # Extract and format the value
                if isinstance(day_data, dict):
                    raw_value = day_data.get('hours', '')
                    
                    # Format if it's a number (minutes)
                    if isinstance(raw_value, (int, float)) and raw_value > 0:
                        cell_val = format_minutes(raw_value)  # 510 -> 8:30
                    elif isinstance(raw_value, str):
                        cell_val = raw_value  # "CO", "CM", "EF", or ""
                    else:
                        cell_val = ''
                else:
                    cell_val = ''
                
                # Add to PDF cell
                data_row.append(Paragraph(str(cell_val if cell_val is not None else ''), body_cell_style))


            # Append totals and counts    
            data_row.extend([
                Paragraph(total_formatted, body_cell_style),
                Paragraph(str(co_days), body_cell_style),
                Paragraph(str(cm_days), body_cell_style),
                Paragraph(str(ef_days), body_cell_style),
                Paragraph(str(meal_tickets), body_cell_style),
            ])

            table_data.append(data_row)

        # 8. Define Column Widths
        col_widths = [18, 95, 25]
        day_col_width = max(15.5, (520 / num_days))
        col_widths.extend([day_col_width] * num_days)
        col_widths.extend([32, 22, 22, 22, 26])

        # 9. Style Table Grid & Weekend Highlights
        t_style = [
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#666666')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f3f5')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 1),
            ('RIGHTPADDING', (0, 0), (-1, -1), 1),
        ]

        # Light gray background for weekends
        for idx_d, day_int in enumerate(month_days):
            weekday = calendar.weekday(year, month, day_int)
            if weekday in (5, 6):  # Saturday / Sunday
                col_index = 3 + idx_d
                t_style.append(('BACKGROUND', (col_index, 0), (col_index, -1), colors.HexColor('#eaeaea')))

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(t_style))
        elements.append(t)

        # 10. SIGNATURE BLOCKS
        elements.append(Spacer(1, 20))

        sig_data = [
            [
                Paragraph("<b>Sef Paza,</b>", sig_title_style),
                "",
                Paragraph("<b>Director,</b>", sig_title_style)
            ],
            [
                Paragraph("Damian Mihai", sig_name_style),
                "",
                Paragraph("Negutescu Ion Clementin", sig_name_style)
            ],
            [
                Paragraph("Semnatura: _______________________", sig_name_style),
                "",
                Paragraph("Semnatura: _______________________", sig_name_style)
            ]
        ]

        sig_table = Table(sig_data, colWidths=[250, 311, 250])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ]))

        elements.append(sig_table)

        # 11. Render PDF
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Pontaj_{year}_{month:02d}.pdf"'
        return response

class PALActivitiesUploadView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """
    Upload new Activity
    """
    model = Activity
    form_class = PALActivitiesUploadForm
    template_name = 'dashboard/palactivities_upload.html'
    success_url = reverse_lazy('pal')
    paginate_by = 20

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get all activities ordered by code
        activities_list = Activity.objects.all().order_by('code')
        
        # Setup pagination
        paginator = Paginator(activities_list, self.paginate_by)
        page = self.request.GET.get('page')
        activities = paginator.get_page(page)
        
        context['activities'] = activities
        return context

    def get(self, request, *args, **kwargs):
        form = PALActivitiesUploadForm()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request, *args, **kwargs):
        form = PALActivitiesUploadForm(request.POST, request.FILES)
        if form.is_valid():
            if 'file' not in request.FILES:
                messages.error(request, "No file uploaded.")
                return redirect(self.success_url)

            excel_file = request.FILES['file']
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active if wb.active is not None else wb[wb.sheetnames[0]]

                # Normalize headers
                headers = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in sheet[1]]

                if 'code' not in headers or 'name' not in headers:
                    messages.error(request, "Excel file must contain 'code' and 'name' columns.")
                    return redirect(self.success_url)

                code_index = headers.index('code')
                name_index = headers.index('name')

                # transaction
                from django.db import transaction
                with transaction.atomic():
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if not row or all(cell is None for cell in row):
                            continue

                        code_val = row[code_index]
                        name_val = sanitize_romanian(row[name_index])

                        if code_val:
                            # This handles both Creating and Updating
                            Activity.objects.update_or_create(
                                code=code_val,
                                defaults={'name': name_val}
                            )
                
                messages.success(request, "Activities uploaded and synced successfully.")
            except Exception as e:
                messages.error(request, f"Error processing Excel file: {e}")
            return redirect(self.success_url)
        else:
            messages.error(request, "Invalid form submission.")
            return render(request, self.template_name, {'form': form})


class PALActivityCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """
    Create view for Activity
    """
    model = Activity
    template_name = 'dashboard/pal_activity_create.html'
    success_url = reverse_lazy('pal')
    form_class = PALActivityForm

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser


class PALActivityUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """
    Update view for Activity
    """
    model = Activity
    template_name = 'dashboard/pal_activity_edit.html'
    success_url = reverse_lazy('pal')
    form_class = PALActivityForm

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser


class PALActivityDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    Delete view for Activity
    """
    model = Activity
    template_name = 'dashboard/pal_activity_delete.html'
    success_url = reverse_lazy('pal')

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activity'] = self.get_object()
        return context


# the activity program view
def activity_program(request):
    template = "activities/activity-program.html"

    context = {}
    return render(request, template, context)


def get_total_hours_qs(queryset):
    """
    Helper to calculate total hours from start_time and end_time at DB level.
    This assumes end_time and start_time are on the same day.
    """
    return queryset.annotate(
        duration=ExpressionWrapper(
            (F('end_time') - F('start_time')),
            output_field=FloatField()
        )
    ).aggregate(
        # Duration is returned in microseconds, 
        # so we divide by 3,600,000,000 to get hours.
        total=Sum(F('duration')) / 3600000000.0
    )['total'] or 0

def worked_hours_per_member(request):
    today = timezone.now()
    team_members = CustomUser.objects.filter(is_active=True)
    data = []

    for member in team_members:
        qs = Timesheet.objects.filter(
            user=member, 
            date__year=today.year, 
            date__month=today.month
        )
        total_hours = get_total_hours_qs(qs)

        data.append({
            'name': member.get_full_name() or member.username,
            'hours': round(float(total_hours), 1)
        })

    return JsonResponse(data, safe=False)

def yearly_statistics(request):
    current_year = timezone.now().year
    months_data = []
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    for month in range(1, 13):
        month_qs = Timesheet.objects.filter(date__year=current_year, date__month=month)
        
        # Calculate totals
        total_worked = get_total_hours_qs(month_qs)
        
        # For counts, we can still use Count
        stats = month_qs.aggregate(
            holidays=Count('id', filter=Q(description__icontains="holiday")), # Adjust filter if needed
            sick_leaves=Count('id', filter=Q(description__icontains="sick")), # Adjust filter if needed
        )

        months_data.append({
            'month': month_names[month-1],
            'worked_hours': round(float(total_worked), 1),
            'holidays': stats['holidays'],
            'sick_leaves': stats['sick_leaves'],
            'weekend_hours': 0 
        })

    return JsonResponse(months_data, safe=False)


class ActivityProgramCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """
    Create view for Activity Program
    """
    model = ActivityProgram
    form_class = ActivityProgramForm
    template_name = 'activities/activity_program_create.html'
    success_url = reverse_lazy('activity_program_list')  # or PDF generation page

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser


class ActivityProgramListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    List view for Activity Programs
    """
    model = ActivityProgram
    template_name = 'activities/activity_program_list.html'
    context_object_name = 'activity_programs'

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_queryset(self):
        return ActivityProgram.objects.filter(user=self.request.user).order_by('-registration_date')


class ActivityProgramUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """
    Update view for Activity Program
    """
    model = ActivityProgram
    form_class = ActivityProgramForm
    template_name = 'activities/activity_program_edit.html'
    success_url = reverse_lazy('activity_program_list')

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser


class ActivityProgramDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    Delete view for Activity Program
    """
    model = ActivityProgram
    template_name = 'activities/activity_program_delete.html'
    success_url = reverse_lazy('activity_program_list')

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activity_program'] = self.get_object()
        return context

class FundsSourceListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    List view for Funds Source
    """
    model = FundsSource
    template_name = 'dashboard/funds_source.html'
    context_object_name = 'funds'

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
class NewFundsSourceView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """
    Create view for new Funds Source
    """
    model = FundsSource
    form_class = FundsSourceForm
    template_name = 'dashboard/new_funds_source.html'
    success_url = reverse_lazy('funds_source')

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
