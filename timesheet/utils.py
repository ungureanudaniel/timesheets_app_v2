# calendarapp/utils.py
from calendar import HTMLCalendar
from .models import Event, Activity
import openpyxl


class Calendar(HTMLCalendar):
    def __init__(self, year=None, month=None):
        self.year = year
        self.month = month
        super(Calendar, self).__init__()

    # formats a day as a td filter events by day
    def formatday(self, day, events):
        events_per_day = events.filter(start_time__day=day)
        d = ""
        for event in events_per_day:
            d += f"<li> {event.get_html_url} </li>"
        if day != 0:
            return f"<td><span class='date'>{day}</span><ul> {d} </ul></td>"
        return "<td></td>"

    # formats a week as a tr
    def formatweek(self, theweek, events):
        week = ""
        for d, weekday in theweek:
            week += self.formatday(d, events)
        return f"<tr> {week} </tr>"

    # formats a month as a table
    # filter events by year and month
    def formatmonth(self, withyear=True):
        events = Event.objects.filter(
            start_time__year=self.year, start_time__month=self.month
        )
        cal = (
            '<table border="0" cellpadding="0" cellspacing="0" class="calendar">\n'
        )  # noqa
        cal += (
            f"{self.formatmonthname(self.year, self.month, withyear=withyear)}\n"
        )  # noqa
        cal += f"{self.formatweekheader()}\n"
        for week in self.monthdays2calendar(self.year, self.month):
            cal += f"{self.formatweek(week, events)}\n"
        return cal

# Import activities from an Excel file and save them to the database
def import_activities_from_excel(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Iterate through the rows in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        activity_code = row[0]
        activity_description = row[1]

        # Create a new Activity object
        activity = Activity(
            code=activity_code,
            name=activity_description,
        )
        activity.save()
    # Close the workbook
    workbook.close()
    # Return the number of activities imported
    return f"Imported {sheet.max_row - 1} activities from {file_path}"
